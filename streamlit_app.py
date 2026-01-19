import io
import json
import platform
import tempfile
import time
import traceback
import uuid
import urllib.error
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path
import sys

import streamlit as st
from openpyxl import load_workbook
import yaml

# --- Robust module import ---
# On Streamlit Cloud the working directory / import path can vary depending
# on how the app is configured (root vs subfolder). We auto-discover the
# directory containing `turni_generator.py` and add it to sys.path.
_APP_DIR = Path(__file__).resolve().parent
for _cand in (_APP_DIR, _APP_DIR / "src", _APP_DIR / "app"):
    try:
        if (_cand / "turni_generator.py").exists() and str(_cand) not in sys.path:
            sys.path.insert(0, str(_cand))
    except Exception:
        pass

try:
    import turni_generator as tg
except Exception as e:
    # Show a readable error in the app UI (and in Streamlit logs)
    st.error(
        "Errore durante l'import di 'turni_generator.py'. "
        "Controlla che il file sia nella repo (stessa cartella o /src) e che "
        "le dipendenze in requirements.txt siano installabili."
    )
    st.exception(e)
    st.stop()

# Export the functions we use below (keeps the rest of the file unchanged)
create_month_template_xlsx = getattr(tg, "create_month_template_xlsx", None)
generate_schedule = getattr(tg, "generate_schedule", None)
extract_carryover_from_output_xlsx = getattr(tg, "extract_carryover_from_output_xlsx", None)
collect_doctors = getattr(tg, "collect_doctors", None)

_missing = [
    n
    for n, v in {
        "create_month_template_xlsx": create_month_template_xlsx,
        "generate_schedule": generate_schedule,
        "extract_carryover_from_output_xlsx": extract_carryover_from_output_xlsx,
        "collect_doctors": collect_doctors,
    }.items()
    if v is None
]
if _missing:
    st.error(
        "'turni_generator.py' e' stato importato, ma mancano funzioni richieste: "
        + ", ".join(_missing)
        + ".\nVerifica di aver fatto push della versione corretta di turni_generator.py."
    )
    st.stop()


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _get_session_id() -> str:
    # Stable within a single browser session (reruns included)
    if "_session_id" not in st.session_state:
        st.session_state["_session_id"] = str(uuid.uuid4())
    return st.session_state["_session_id"]


def _summarize_stats(stats: dict | None) -> dict:
    """Return a compact summary suitable for audit logs."""
    if not isinstance(stats, dict):
        return {"status": "UNKNOWN"}

    months = stats.get("months") or {}
    month_summary: dict[str, dict] = {}
    greedy_months: list[str] = []
    infeasible_months: list[str] = []

    for k, v in months.items():
        if not isinstance(v, dict):
            month_summary[k] = {"status": str(v)}
            continue
        st_m = str(v.get("status", "")).upper()
        se = v.get("solver_error")
        if se:
            greedy_months.append(k)
        if "INFEAS" in st_m:
            infeasible_months.append(k)
        month_summary[k] = {
            "status": v.get("status"),
            "solver_error": (str(se)[:400] if se else None),
            "autorelax": v.get("autorelax"),
        }

    return {
        "status": stats.get("status"),
        "greedy_months": greedy_months,
        "infeasible_months": infeasible_months,
        "months": month_summary,
    }


def _github_audit_log(event: dict) -> tuple[bool, str]:
    """Append an audit event as a comment to a GitHub Issue.

    Configure in Streamlit Secrets (Manage app → Settings → Secrets):

    [github]
    token = "..."
    repo  = "owner/repo"
    issue = 1

    Token needs permission to write Issue comments.
    """
    try:
        gh = st.secrets.get("github", {})
        token = gh.get("token") or st.secrets.get("GITHUB_TOKEN")
        repo = gh.get("repo")
        issue = gh.get("issue")
        if not (token and repo and issue):
            return False, "github audit log not configured"

        url = f"https://api.github.com/repos/{repo}/issues/{int(issue)}/comments"

        # Human-friendly first line + JSON payload
        headline = (
            f"{event.get('result', 'unknown').upper()} | "
            f"{event.get('year')}-{int(event.get('month') or 0):02d} | "
            f"template={event.get('template_mode')} | "
            f"sheet={event.get('sheet_name_used')} | "
            f"operator={event.get('operator') or '-'}"
        )
        body = {
            "body": headline
            + "\n\n```json\n"
            + json.dumps(event, ensure_ascii=False)
            + "\n```"
        }

        data = json.dumps(body).encode("utf-8")
        req = urllib.request.Request(
            url,
            data=data,
            method="POST",
            headers={
                "Authorization": f"Bearer {token}",
                "Accept": "application/vnd.github+json",
                "User-Agent": "turni-autogen-streamlit",
                "Content-Type": "application/json",
            },
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return True, f"ok ({resp.status})"
    except urllib.error.HTTPError as e:
        try:
            detail = e.read().decode("utf-8", errors="ignore")[:500]
        except Exception:
            detail = ""
        return False, f"HTTPError {getattr(e, 'code', '')} {detail}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _month_label_it(month: int) -> str:
    names = {
        1: "Gennaio",
        2: "Febbraio",
        3: "Marzo",
        4: "Aprile",
        5: "Maggio",
        6: "Giugno",
        7: "Luglio",
        8: "Agosto",
        9: "Settembre",
        10: "Ottobre",
        11: "Novembre",
        12: "Dicembre",
    }
    return names.get(int(month), str(month))


st.set_page_config(page_title="Turni Autogenerator", layout="wide")

st.title("Turni Autogenerator – versione web (Streamlit)")

st.markdown(
    """
Questa versione gira **senza Tkinter** (Streamlit Cloud non supporta GUI desktop).

**Default:** selezioni **mese** + **anno** e il programma **genera automaticamente il template Excel** (date + intestazioni colonne dalla YAML).

**Avanzato:** puoi caricare un tuo *Template Excel*.
"""
)

# --- Selezione mese/anno ---
col_a, col_b, col_c = st.columns([1, 1, 2])
with col_a:
    year = st.number_input("Anno", min_value=2024, max_value=2035, value=2026, step=1)
with col_b:
    month = st.selectbox("Mese", options=list(range(1, 13)), format_func=_month_label_it, index=1)
with col_c:
    operator_tag = st.text_input(
        "Operatore (opzionale)",
        value="",
        help="Nome/codice di chi sta usando l'app (per audit log).",
    )

st.divider()

# --- Regole ---
col1, col2 = st.columns(2)
with col1:
    use_repo_rules = st.checkbox("Usa Regole_Turni.yml del repo", value=True)

with col2:
    rules_up = None
    if not use_repo_rules:
        rules_up = st.file_uploader("Regole (.yml/.yaml)", type=["yml", "yaml"], accept_multiple_files=False)

# --- Indisponibilità (lasciata com'è) ---
unav_up = st.file_uploader(
    "Indisponibilità (opzionale: .xlsx/.csv/.tsv)",
    type=["xlsx", "xls", "csv", "tsv"],
    accept_multiple_files=False,
)

prev_out_up = st.file_uploader(
    "Output mese precedente (opzionale: .xlsx, per vincoli inter-mese su NOTTI)",
    type=["xlsx"],
    accept_multiple_files=False,
)

prev_sheet_name_from_upload: str | None = None
if prev_out_up is not None:
    try:
        wb_prev = load_workbook(io.BytesIO(prev_out_up.getvalue()), read_only=True, data_only=True)
        sheets_prev = wb_prev.sheetnames
    except Exception:
        sheets_prev = []
    if sheets_prev:
        st.caption("Fogli trovati nel file precedente: " + ", ".join(sheets_prev))
        optp = st.selectbox(
            "Seleziona foglio (file precedente)",
            options=["(foglio attivo / primo foglio)"] + sheets_prev,
            index=0,
        )
        prev_sheet_name_from_upload = None if optp.startswith("(") else optp
    else:
        st.warning("Non riesco a leggere i fogli del file precedente: userò il primo foglio.")
        prev_sheet_name_from_upload = None


# --- Carryover manuale (quando non hai l'output precedente) ---
# Best-effort list of doctors from rules (to reduce typos)
_doctors_ui = []
try:
    if use_repo_rules:
        _rp = Path(__file__).with_name("Regole_Turni.yml")
        if _rp.exists():
            _cfg_ui = yaml.safe_load(_rp.read_text(encoding="utf-8"))
            if isinstance(_cfg_ui, dict):
                _doctors_ui = [d for d in collect_doctors(_cfg_ui) if d and d != "Recupero"]
    else:
        if rules_up is not None:
            _cfg_ui = yaml.safe_load(rules_up.getvalue())
            if isinstance(_cfg_ui, dict):
                _doctors_ui = [d for d in collect_doctors(_cfg_ui) if d and d != "Recupero"]
except Exception:
    _doctors_ui = []

manual_block_first_selected = st.multiselect(
    "Se non hai il file precedente: seleziona chi ha fatto NOTTE l’ultimo giorno del mese precedente (blocco su Giorno 1)",
    options=_doctors_ui,
    default=[],
    help="Questi medici verranno esclusi da tutti i compiti nel Giorno 1 del mese che stai generando (vincolo inter-mese).",
)

manual_block_first_text = st.text_input(
    "Oppure scrivi altri nomi (separati da virgola)",
    value="",
    help="Utile se un nome non compare nell'elenco o se hai un refuso nel YAML da correggere.",
)
st.divider()

# --- Template: default auto, upload in Avanzate ---
with st.expander("Avanzate: template Excel", expanded=False):
    use_custom_template = st.checkbox(
        "Usa template personalizzato (upload)",
        value=False,
        help="Se disattivato, il template viene creato automaticamente da mese/anno + regole.",
    )

    template_up = None
    sheet_name_from_upload: str | None = None

    if use_custom_template:
        template_up = st.file_uploader("Template turni (.xlsx)", type=["xlsx"], accept_multiple_files=False)

        if template_up is not None:
            # Fogli reali dal template → dropdown (niente input libero)
            try:
                wb_tmp = load_workbook(io.BytesIO(template_up.getvalue()), read_only=True, data_only=True)
                sheets = wb_tmp.sheetnames
            except Exception:
                sheets = []

            if sheets:
                st.caption("Fogli trovati nel template: " + ", ".join(sheets))
                opt = st.selectbox(
                    "Seleziona foglio",
                    options=["(foglio attivo / primo foglio)"] + sheets,
                    index=0,
                )
                sheet_name_from_upload = None if opt.startswith("(") else opt
            else:
                st.warning("Non riesco a leggere i fogli: verrà usato il foglio attivo (primo foglio).")
                sheet_name_from_upload = None
        else:
            st.info("Carica un template .xlsx per abilitarne la selezione foglio.")

    else:
        st.caption(
            "Template auto: verrà creato un Excel con le date del mese e le intestazioni colonne dalla YAML. "
            "Il nome del foglio sarà generato automaticamente."
        )

run_btn = st.button("Genera turni", type="primary")

if run_btn:
    # --- Validate rules input ---
    if not use_repo_rules and rules_up is None:
        st.error("Hai disattivato 'Usa Regole_Turni.yml del repo': carica un file **Regole (.yml/.yaml)**.")
        st.stop()

    # --- Validate template if custom ---
    if use_custom_template and template_up is None:
        st.error("Hai scelto 'Usa template personalizzato': carica un **Template turni (.xlsx)**.")
        st.stop()

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)

        # Save rules
        if use_repo_rules:
            repo_rules = Path(__file__).with_name("Regole_Turni.yml")
            if not repo_rules.exists():
                st.error("Non trovo 'Regole_Turni.yml' nel repo. Carica un file regole manualmente.")
                st.stop()
            rules_path = td / "Regole_Turni.yml"
            rules_path.write_bytes(repo_rules.read_bytes())
            rules_source = "repo"
        else:
            rules_path = td / "Regole_Turni.yml"
            rules_path.write_bytes(rules_up.getvalue())
            rules_source = getattr(rules_up, "name", "upload")

        # Prepare template
        if use_custom_template:
            template_path = td / "template.xlsx"
            template_path.write_bytes(template_up.getvalue())
            template_mode = "upload"
            sheet_name_used = sheet_name_from_upload
            template_filename = getattr(template_up, "name", None)
            template_bytes = len(template_up.getvalue()) if template_up is not None else None
        else:
            month_name = _month_label_it(int(month)).upper()
            sheet_auto = f"GUARDIE_{month_name}_{int(year)}"
            template_path = td / f"template_{int(year)}_{int(month):02d}.xlsx"
            create_month_template_xlsx(
                rules_yml=rules_path,
                year=int(year),
                month=int(month),
                out_path=template_path,
                sheet_name=sheet_auto,
            )
            template_mode = "auto"
            sheet_name_used = sheet_auto
            template_filename = template_path.name
            template_bytes = template_path.stat().st_size if template_path.exists() else None

        # Save unavailability (optional)
        unav_path = None
        if unav_up is not None:
            unav_path = td / f"unavailability.{unav_up.name.split('.')[-1]}"
            unav_path.write_bytes(unav_up.getvalue())

        out_path = td / f"turni_{int(year)}_{int(month):02d}.xlsx"

        carryover_by_month = None
        if prev_out_up is not None:
            prev_path = td / f"prev_output_{prev_out_up.name}"
            prev_path.write_bytes(prev_out_up.getvalue())
            carry = extract_carryover_from_output_xlsx(prev_path, rules_path, sheet_name=prev_sheet_name_from_upload)
            # Safety: apply the automatic "block day 1" only when the previous file really ends
            # the day before the month we are generating.
            try:
                target_first = datetime(int(year), int(month), 1).date()
                last_date_str = (carry or {}).get("last_date")
                last_date = datetime.fromisoformat(last_date_str).date() if last_date_str else None
                if last_date and last_date != (target_first - timedelta(days=1)):
                    if (carry or {}).get("block_all_on_first_day"):
                        carry["block_all_on_first_day"] = []
                    st.warning(
                        f"Il file precedente sembra NON contiguo al mese richiesto: ultima data letta = {last_date}. "
                        f"Per sicurezza disattivo il blocco automatico sul 1° giorno. "
                        f"(Puoi sempre usare il campo manuale qui sopra.)"
                    )
            except Exception:
                pass
            carryover_by_month = {(int(year), int(month)): carry}
            st.info(carry.get('note') or "Carryover letto dal mese precedente.")

        # Manual override/integrazione: blocco giorno 1 per NOTTE ultimo giorno mese precedente
        manual_names = []
        # from multiselect
        manual_names.extend([x.strip() for x in (manual_block_first_selected or []) if str(x).strip()])
        # from free-text fallback
        manual_names.extend([x.strip() for x in (manual_block_first_text or '').split(',') if x.strip()])
        # de-duplicate, preserve order
        seen = set()
        manual_names = [n for n in manual_names if not (n in seen or seen.add(n))]
        if manual_names:
            if carryover_by_month is None:
                carryover_by_month = {(int(year), int(month)): {'recent_nights': [], 'block_all_on_first_day': manual_names, 'note': 'Carryover manuale (solo blocco giorno 1)'}}
            else:
                key = (int(year), int(month))
                carry = carryover_by_month.get(key) or {}
                cur = list(carry.get('block_all_on_first_day') or [])
                for n in manual_names:
                    if n not in cur:
                        cur.append(n)
                carry['block_all_on_first_day'] = cur
                carryover_by_month[key] = carry
            st.info('Carryover manuale: blocco su giorno 1 per: ' + ', '.join(manual_names))


        run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_id = _get_session_id()
        base_event = {
            "ts_utc": _utc_now_iso(),
            "run_id": run_id,
            "session_id": session_id,
            "operator": operator_tag.strip() or None,
            "year": int(year),
            "month": int(month),
            "template_mode": template_mode,
            "template_filename": template_filename,
            "template_bytes": template_bytes,
            "sheet_name_used": sheet_name_used,
            "rules_source": rules_source,
            "unavailability_filename": getattr(unav_up, "name", None) if unav_up is not None else None,
            "unavailability_bytes": len(unav_up.getvalue()) if unav_up is not None else None,
            "python": platform.python_version(),
        }

        t0 = time.time()
        try:
            with st.spinner("Calcolo in corso…"):
                stats, log_path = generate_schedule(
                    template_xlsx=template_path,
                    rules_yml=rules_path,
                    out_xlsx=out_path,
                    unavailability_path=unav_path,
                    sheet_name=sheet_name_used,
                    carryover_by_month=carryover_by_month,
                )

            duration = round(time.time() - t0, 3)
            event = {
                **base_event,
                "result": "ok",
                "duration_s": duration,
                "stats": _summarize_stats(stats),
            }
            ok, msg = _github_audit_log(event)
            if not ok:
                st.caption(f"Audit log GitHub non scritto: {msg}")
        except Exception as e:
            duration = round(time.time() - t0, 3)
            event = {
                **base_event,
                "result": "error",
                "duration_s": duration,
                "error_type": type(e).__name__,
                "error": str(e),
                "traceback": traceback.format_exc()[:8000],
            }
            _github_audit_log(event)
            st.error("Errore durante la generazione dei turni.")
            st.code(event["traceback"])
            st.stop()

        # --- Results ---
        st.success("Turni generati.")

        # Show solver summary
        status = (stats or {}).get("status", "")
        st.subheader("Esito solver")
        st.write(f"**Status:** {status}")

        months_stats = (stats or {}).get("months") or {}
        greedy_months = [k for k, v in months_stats.items() if isinstance(v, dict) and v.get("solver_error")]
        if greedy_months:
            st.warning(
                "OR-Tools non disponibile o schedule infeasible per: "
                + ", ".join(greedy_months)
                + ". In quei mesi è stato usato il greedy."
            )

        # Download output
        st.download_button(
            label="Scarica Excel generato",
            data=out_path.read_bytes(),
            file_name=out_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Offer generated template download when in auto mode
        if template_mode == "auto" and template_path.exists():
            st.download_button(
                label="Scarica template auto (debug)",
                data=template_path.read_bytes(),
                file_name=template_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # Show log (if any)
        if log_path and Path(log_path).exists():
            st.subheader("Log")
            try:
                st.code(Path(log_path).read_text(encoding="utf-8", errors="ignore"))
            except Exception:
                st.code(Path(log_path).read_text(errors="ignore"))
